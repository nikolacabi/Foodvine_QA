import random
import openpyxl


# l is last city in list to be considered
def gencity():

    wb = openpyxl.load_workbook('/storage/data/Cities.xlsx', read_only=True)
    ws = wb['Details']

    row_count = ws.max_row

    n = random.randint(1, row_count - 1)

    city_id = ws['A' + str(n + 1)].value
    city_name = ws['B' + str(n + 1)].value
    postal_code = ws['C' + str(n + 1)].value
    latitude = ws['D' + str(n + 1)].value
    longitude = ws['E' + str(n + 1)].value
    farmer_market = ws['F' + str(n + 1)].value
    recipes = ws['G' + str(n + 1)].value
    live_classes = ws['H' + str(n + 1)].value
    order_from = ws['I' + str(n + 1)].value

    city = [city_id, city_name, postal_code, latitude, longitude, farmer_market, recipes, live_classes, order_from]

    wb.close()

    return city

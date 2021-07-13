from datetime import datetime
import openpyxl
import random


USER_SHEET_LOC = 'storage/data/Users.xlsx'


# Function that generates unsername, email and password and returns them
def genuser():

    un = 'testnikola' + datetime.now().strftime("%d%m%Y%H%M%S")
    email = un + "@gmail.com"
    pw = '1234Test'

    wb = openpyxl.load_workbook(USER_SHEET_LOC, read_only=True)
    ws = wb['Users']
    row_count = ws.max_row      # row count can be read only in read_only=True so this is here and not in below part
    wb.close()

    return un, email, pw




def saveuser(un, email, pw):

    wb = openpyxl.load_workbook(USER_SHEET_LOC, read_only=True)
    ws = wb['Users']
    row_count = ws.max_row      # row count can be read only in read_only=True so this is here and not in below part
    wb.close()

    wb = openpyxl.load_workbook(USER_SHEET_LOC)
    ws = wb['Users']

    ws['A' + str(row_count + 1)] = un
    ws['B' + str(row_count + 1)] = email
    ws['C' + str(row_count + 1)] = pw

    wb.save(USER_SHEET_LOC)
    wb.close()


# Function that returns random user credentials
def getuser():

    wb = openpyxl.load_workbook(USER_SHEET_LOC, read_only=True)
    ws = wb['Users']
    row_count = ws.max_row

    n = random.randint(1, row_count - 1)

    un = ws['A' + str(n + 1)].value
    email = ws['B' + str(n + 1)].value
    pw = ws['C' + str(n + 1)].value

    wb.close()

    return un, email, pw




def changepw(un, email, new_pw):

    wb = openpyxl.load_workbook(USER_SHEET_LOC, read_only=True)
    ws = wb['Users']
    row_count = ws.max_row
    wb.close()

    wb = openpyxl.load_workbook(USER_SHEET_LOC)
    ws = wb['Users']

    for i in range(0, row_count):
        if un == ws['A' + str(i + 1)].value and email == ws['B' + str(i + 1)].value:
            ws['C' + str(i + 1)] = new_pw

    wb.save(USER_SHEET_LOC)
    wb.close()